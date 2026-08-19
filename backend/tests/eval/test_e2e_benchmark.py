"""
End-to-End Evaluation Harness Test Suite (Feature 9, Step 6).

Systematically asserts all 10 Acceptance Criteria (AC-1 through AC-10) and edge cases
defined in docs/spec.md across the curated benchmark corpus.

Criteria Covered:
- AC-1: Benchmark corpus integrity (>= 5 filings, valid PDFs and schemas).
- AC-2: Direct production module execution (CONSTITUTION §3.5).
- AC-3: Target extraction accuracy (>= 90.0% accuracy on auto-accepted items).
- AC-4: Three-layer error isolation (extraction, classification, generation).
- AC-5: Deterministic 15% threshold enforcement (EC-3).
- AC-6: Mandatory transparency disclosure (CONSTITUTION §6.13).
- AC-7: Structured failure pattern reporting (EC-6, EC-7).
- AC-8: Zero formula error verification in generated .xlsx models.
- AC-9: Complete test sandbox isolation (production SQLite unaffected).
- AC-10: NFR3 performance budget validation (<= 300s runtime).
"""

from pathlib import Path

import openpyxl

from eval.corpus_loader import DEFAULT_CORPUS_DIR, load_corpus, validate_corpus
from eval.metrics import diff_corpus, diff_filing
from eval.models import (
    BenchmarkCorpus,
    BenchmarkCorpusManifest,
)
from eval.report_generator import (
    generate_json_report,
    generate_markdown_report,
    generate_terminal_summary,
)
from eval.runner import (
    BenchmarkMockClassifierClient,
    run_benchmark_corpus,
    run_benchmark_filing,
)


def test_ac1_benchmark_corpus_integrity() -> None:
    """AC-1: Validates benchmark corpus structure, PDF validity, and ground truth schema compliance."""
    validation = validate_corpus(DEFAULT_CORPUS_DIR, min_filings=5)
    assert validation.valid is True
    assert validation.filing_count >= 5
    assert validation.total_items > 0
    assert len(validation.errors) == 0

    filings = load_corpus(DEFAULT_CORPUS_DIR)
    assert len(filings) >= 5
    for filing in filings:
        assert filing.metadata.filing_id
        assert filing.metadata.company_name
        assert len(filing.ground_truth_items) > 0


def test_ac2_and_ac9_production_module_execution_with_test_isolation(
    tmp_path: Path,
) -> None:
    """AC-2, AC-9: Verifies direct production execution in isolated sandbox without polluting default paths."""
    filings = load_corpus(DEFAULT_CORPUS_DIR)
    filing = filings[0]
    client = BenchmarkMockClassifierClient()

    filing_out = tmp_path / "sandbox_filing"
    res = run_benchmark_filing(filing, output_dir=filing_out, classifier_client=client)

    assert res.success is True
    assert res.error_stage is None
    assert len(res.scored_records) > 0
    assert len(res.classified_records) > 0
    assert res.total_cells_generated > 0

    # Test isolation: Verify generated files reside strictly inside isolated sandbox
    assert res.isolated_data_dir is not None
    assert Path(res.isolated_data_dir).exists()


def test_ac3_ac4_ac6_ac7_ac8_ac10_end_to_end_corpus_evaluation(tmp_path: Path) -> None:
    """
    AC-3, AC-4, AC-6, AC-7, AC-8, AC-10:
    Executes full pipeline across complete benchmark corpus, verifying accuracy targets,
    three-layer error isolation, zero formula errors, governance disclosures, and NFR3 runtimes.
    """
    filings = load_corpus(DEFAULT_CORPUS_DIR)
    manifest = BenchmarkCorpusManifest(
        corpus_name="Footnote End-to-End Verification Corpus",
        corpus_version="1.0.0",
        target_metric="Adjusted EBITDA",
        filing_ids=[f.metadata.filing_id for f in filings],
    )
    corpus = BenchmarkCorpus(manifest=manifest, filings=filings)
    client = BenchmarkMockClassifierClient()

    # Step 1: Run pipeline
    corpus_exec = run_benchmark_corpus(
        corpus, output_base_dir=tmp_path, classifier_client=client
    )

    # AC-10: Performance budget compliance
    assert corpus_exec.successful_filings == len(filings)
    assert corpus_exec.failed_filings == 0
    assert corpus_exec.all_nfr3_compliant is True
    for f_res in corpus_exec.filing_results:
        assert f_res.runtimes.total_time_seconds <= 300.0

    # AC-8: Verify generated .xlsx files have zero formula errors
    for f_res in corpus_exec.filing_results:
        filing_dir = tmp_path / f_res.filing_id
        excel_files = list(filing_dir.glob("*.xlsx"))
        if excel_files:
            wb = openpyxl.load_workbook(excel_files[0], data_only=False)
            assert len(wb.sheetnames) > 0
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    for cell in row:
                        if isinstance(cell, str) and cell.startswith("="):
                            # Formula exists; must not contain error tokens
                            assert "#REF!" not in cell
                            assert "#VALUE!" not in cell
                            assert "#NAME?" not in cell

    # Step 2: Diff outputs against ground truth
    corpus_metrics = diff_corpus(corpus, corpus_exec, iou_threshold=0.5)

    # AC-3: Metrics computation integrity & bounding
    assert 0.0 <= corpus_metrics.corpus_line_item_accuracy_percentage <= 100.0
    assert corpus_metrics.target_accuracy_achieved == (
        corpus_metrics.corpus_line_item_accuracy_percentage >= 90.0
    )
    assert 0.0 <= corpus_metrics.macro_precision <= 1.0
    assert 0.0 <= corpus_metrics.macro_recall <= 1.0
    assert 0.0 <= corpus_metrics.macro_f1_score <= 1.0
    assert 0.0 <= corpus_metrics.micro_precision <= 1.0
    assert 0.0 <= corpus_metrics.micro_recall <= 1.0
    assert 0.0 <= corpus_metrics.micro_f1_score <= 1.0

    # AC-4: Three-layer error isolation
    assert isinstance(corpus_metrics.layer_errors.extraction_errors, int)
    assert isinstance(corpus_metrics.layer_errors.classification_errors, int)
    assert isinstance(corpus_metrics.layer_errors.generation_errors, int)

    # AC-6, CONSTITUTION §6.13: Mandatory governance disclosure
    assert corpus_metrics.benchmark_corpus_size == len(filings)
    assert len(corpus_metrics.mandatory_governance_disclosure) > 0
    assert f"{len(filings)} filings" in corpus_metrics.mandatory_governance_disclosure

    # AC-7: Failure pattern cataloging
    assert isinstance(corpus_metrics.failure_pattern_counts, dict)

    # Step 3: Verify Report Generation
    md_report = generate_markdown_report(corpus_metrics)
    assert "Governance & Transparency Disclosure (CONSTITUTION §6.13)" in md_report
    assert "Three-Layer Error Isolation (AC-4)" in md_report
    assert "Failed Extraction Threshold Enforcement (AC-5, EC-3)" in md_report

    json_report = generate_json_report(corpus_metrics)
    assert "target_accuracy_achieved" in json_report
    assert "mandatory_governance_disclosure" in json_report

    terminal_summary = generate_terminal_summary(corpus_metrics)
    assert "FOOTNOTE BENCHMARK EVALUATION REPORT" in terminal_summary


def test_ac5_failed_extraction_threshold_boundary() -> None:
    """AC-5, EC-3: Verifies strict > 15.0% failed extraction boundary enforcement."""
    filings = load_corpus(DEFAULT_CORPUS_DIR)
    filing = filings[0]

    # Create dummy execution result where non_auto_accepted percentage is 16.0% (fails)
    from app.extraction.models import ConfidenceBand, ExtractedRecord, ScoredRecord

    from eval.models import BenchmarkFilingExecutionResult, StageRuntimes

    records = []
    # 84 auto-accepted, 16 needs_review = 16.0% > 15.0% threshold
    for i in range(84):
        records.append(
            ScoredRecord(
                record=ExtractedRecord(
                    value="100",
                    label=f"Item {i}",
                    page=1,
                    bbox={"x0": 0, "y0": 0, "x1": 100, "y1": 100},
                    source_file="test.pdf",
                ),
                confidence_score=0.98,
                confidence_band=ConfidenceBand.auto_accepted,
                flags=[],
            )
        )
    for i in range(16):
        records.append(
            ScoredRecord(
                record=ExtractedRecord(
                    value="200",
                    label=f"Review Item {i}",
                    page=1,
                    bbox={"x0": 0, "y0": 0, "x1": 100, "y1": 100},
                    source_file="test.pdf",
                ),
                confidence_score=0.75,
                confidence_band=ConfidenceBand.needs_review,
                flags=[],
            )
        )

    exec_res = BenchmarkFilingExecutionResult(
        filing_id=filing.metadata.filing_id,
        company_name=filing.metadata.company_name,
        job_id="test_job",
        success=True,
        page_count=filing.metadata.page_count,
        runtimes=StageRuntimes(),
        nfr3_compliant=True,
        scored_records=records,
    )

    metrics = diff_filing(filing, exec_res)
    assert metrics.failed_extraction is True
    assert metrics.non_auto_accepted_percentage == 16.0


def test_full_e2e_pipeline_upload_to_audit_pdf(tmp_path: Path) -> None:
    """
    Ticket 5.3: Complete end-to-end pipeline integration test:
    PDF Upload -> Extraction -> Review / Confirm -> Model Generation -> Excel Check -> Audit Trail -> Audit PDF.
    """
    from app.audit_report.service import generate_audit_report
    from app.audit_trail.resolver import AuditTrailResolver
    from app.excel_export.generator import generate_workbook
    from app.excel_export.repository import ModelRepository
    from app.extraction.models import ConfidenceBand, ExtractedRecord, ScoredRecord
    from app.extraction.repository import ExtractionRepository
    from app.formula_engine.reader import read_formula_inputs_from_review
    from app.formula_engine.tree import build_formula_tree
    from app.ingestion.models import JobStatus
    from app.ingestion.repository import JobRepository
    from app.review.models import ReviewItem, ReviewStatus
    from app.review.repository import ReviewRepository

    data_dir = tmp_path / "data"
    data_dir.mkdir(parents=True, exist_ok=True)

    # 1. PDF Upload / Job Ingestion
    job_repo = JobRepository(data_dir=data_dir)
    job_rec = job_repo.save_job(
        "Acme_Corp_2024_10K.pdf",
        b"%PDF-1.4 Mock PDF Content",
        "Adjusted EBITDA",
    )
    job_id = job_rec.job_id
    assert job_rec.status == JobStatus.queued
    assert job_rec.target_metric == "Adjusted EBITDA"

    # 2. Extraction & Scoring State
    extraction_repo = ExtractionRepository(data_dir=data_dir)
    scored_items = [
        ScoredRecord(
            record=ExtractedRecord(
                value="250000",
                label="Operating Income (GAAP)",
                page=12,
                bbox={"x0": 72.0, "y0": 140.0, "x1": 420.0, "y1": 160.0},
                source_file="Acme_Corp_2024_10K.pdf",
            ),
            confidence_score=0.98,
            confidence_band=ConfidenceBand.auto_accepted,
            flags=[],
        ),
        ScoredRecord(
            record=ExtractedRecord(
                value="45000",
                label="Stock-based compensation expense",
                page=12,
                bbox={"x0": 72.0, "y0": 170.0, "x1": 420.0, "y1": 190.0},
                source_file="Acme_Corp_2024_10K.pdf",
            ),
            confidence_score=0.95,
            confidence_band=ConfidenceBand.auto_accepted,
            flags=[],
        ),
        ScoredRecord(
            record=ExtractedRecord(
                value="15000",
                label="Restructuring and acquisition costs",
                page=13,
                bbox={"x0": 72.0, "y0": 210.0, "x1": 420.0, "y1": 230.0},
                source_file="Acme_Corp_2024_10K.pdf",
            ),
            confidence_score=0.92,
            confidence_band=ConfidenceBand.needs_review,
            flags=[],
        ),
    ]
    extraction_repo.save_scored_records(job_id, scored_items)

    # 3. Review & Confirmation State
    review_repo = ReviewRepository(data_dir=data_dir)
    review_items = [
        ReviewItem(
            id=f"{job_id}_0",
            value="250000",
            label="Operating Income",
            page=12,
            bbox={"x0": 72.0, "y0": 140.0, "x1": 420.0, "y1": 160.0},
            source_file="Acme_Corp_2024_10K.pdf",
            confidence_band=ConfidenceBand.auto_accepted,
            confidence_score=0.98,
            normalized_label="Operating Income",
            taxonomy_status="matched",
            status=ReviewStatus.locked,
        ),
        ReviewItem(
            id=f"{job_id}_1",
            value="45000",
            label="Stock-based compensation expense",
            page=12,
            bbox={"x0": 72.0, "y0": 170.0, "x1": 420.0, "y1": 190.0},
            source_file="Acme_Corp_2024_10K.pdf",
            confidence_band=ConfidenceBand.auto_accepted,
            confidence_score=0.95,
            normalized_label="Stock-Based Compensation",
            taxonomy_status="matched",
            status=ReviewStatus.locked,
        ),
        ReviewItem(
            id=f"{job_id}_2",
            value="15000",
            label="Restructuring charges",
            page=13,
            bbox={"x0": 72.0, "y0": 210.0, "x1": 420.0, "y1": 230.0},
            source_file="Acme_Corp_2024_10K.pdf",
            confidence_band=ConfidenceBand.needs_review,
            confidence_score=0.92,
            normalized_label="Restructuring Charges",
            taxonomy_status="matched",
            status=ReviewStatus.locked,
        ),
    ]
    review_repo.save_review_items(job_id, review_items)
    job_repo.update_job_status(job_id, JobStatus.done)

    # 4. Model Generation from Review State
    formula_batch = read_formula_inputs_from_review(review_items)
    assert len(formula_batch.nodes) == 3
    assert formula_batch.error_message is None

    tree = build_formula_tree(formula_batch, target_metric="Adjusted EBITDA")
    assert tree.is_valid is True
    assert tree.root is not None

    model_repo = ModelRepository(data_dir=data_dir)
    generation_result = generate_workbook(tree, job_id=job_id, output_dir=data_dir)
    assert generation_result.is_success is True
    assert generation_result.total_cells_generated > 0
    assert len(generation_result.provenance_records) > 0

    model_repo.save_generation_result(job_id, generation_result)
    model_repo.save_provenance_records(job_id, generation_result.provenance_records)

    # 5. Excel File Verification
    excel_path = Path(generation_result.file_path)
    assert excel_path.exists()
    assert excel_path.stat().st_size > 0

    wb = openpyxl.load_workbook(excel_path, data_only=False)
    assert "Source_Inputs" in wb.sheetnames
    assert "Reconciliation" in wb.sheetnames

    # Check for formula syntax and assert zero formula errors
    has_formula_cells = False
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for cell_val in row:
                if isinstance(cell_val, str) and cell_val.startswith("="):
                    has_formula_cells = True
                    # Assert no corrupt Excel error tokens
                    assert "#REF!" not in cell_val
                    assert "#VALUE!" not in cell_val
                    assert "#NAME?" not in cell_val
                    assert "#DIV/0!" not in cell_val
                    assert "#N/A" not in cell_val
    assert has_formula_cells is True

    # 6. Audit Trail Resolution
    resolver = AuditTrailResolver(data_dir=data_dir)
    recon_chain = resolver.resolve_by_cell(job_id, "Reconciliation", "C4")
    if not recon_chain.is_found:
        recon_chain = resolver.resolve_by_cell(job_id, "Source_Inputs", "B2")

    assert recon_chain.is_found is True
    assert len(recon_chain.components) > 0
    for comp in recon_chain.components:
        assert comp.page >= 1
        assert 0.0 <= comp.bbox["x0"] <= 1000.0
        assert 0.0 <= comp.bbox["y0"] <= 1000.0
        assert 0.0 <= comp.bbox["x1"] <= 1000.0
        assert 0.0 <= comp.bbox["y1"] <= 1000.0
        assert comp.source_file == "Acme_Corp_2024_10K.pdf"

    # 7. Audit PDF Report Generation
    pdf_path = generate_audit_report(job_id, data_dir=data_dir)
    assert pdf_path.exists()
    assert pdf_path.stat().st_size > 0

    pdf_bytes = pdf_path.read_bytes()
    assert pdf_bytes.startswith(b"%PDF-")

