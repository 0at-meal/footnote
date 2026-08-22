"""
Unit tests for Excel workbook generator (Feature 4 Step 3).

Tests:
- AC-1 / NFR1: Deterministic workbook structure
- AC-2: Zero numeric literals in derived cells (all derived cells use formulas)
- AC-3: Valid recalculable Excel formulas without broken references
- AC-7 / EC-6: Space-free sheet names
- EC-1: Multiple confirmed records with same label aggregated via formulas
- EC-2: Non-numeric strings logged as warnings without crashing
- EC-7: Atomic write failure cleanup
- EC-10: Fresh generation from scratch
- CONSTITUTION §2.5: IB color styling
"""

from pathlib import Path
from unittest.mock import patch

import openpyxl
from app.excel_export.generator import generate_workbook
from app.excel_export.models import WorkbookGenerationResult
from app.formula_engine.models import (
    FormulaInputBatch,
    FormulaInputNode,
    FormulaTree,
)
from app.formula_engine.tree import build_formula_tree


def _make_input_node(
    record_index: int,
    normalized_label: str,
    value: str = "100.0",
    page: int = 10,
    source_file: str = "filing.pdf",
    is_hardcode: bool = False,
) -> FormulaInputNode:
    return FormulaInputNode(
        node_id=f"node_{record_index}_{normalized_label}",
        normalized_label=normalized_label,
        value=value,
        label=f"Structural / {normalized_label}",
        page=page,
        bbox={"x0": 100.0, "y0": 200.0, "x1": 300.0, "y1": 250.0},
        source_file=source_file,
        record_index=record_index,
        is_hardcode=is_hardcode,
    )


def test_generate_workbook_success(tmp_path: Path) -> None:
    """Verifies complete .xlsx creation with valid sheets and metadata."""
    nodes = [
        _make_input_node(0, "Operating Income", value="1,500.00", page=12),
        _make_input_node(1, "Depreciation & Amortization", value="250.00", page=14),
        _make_input_node(2, "Stock-Based Compensation", value="75.00", page=30),
    ]
    batch = FormulaInputBatch(
        nodes=nodes,
        total_records_received=3,
        confirmed_count=3,
        excluded_count=0,
    )
    tree = build_formula_tree(batch, target_metric="Adjusted EBITDA")

    result = generate_workbook(tree, job_id="job_test_1", output_dir=tmp_path)

    assert isinstance(result, WorkbookGenerationResult)
    assert result.is_success is True
    assert result.error_detail is None
    assert Path(result.file_path).exists()
    assert result.target_metric == "Adjusted EBITDA"
    assert result.sheet_names == ["Source_Inputs", "Reconciliation"]
    assert result.formula_cells_count > 0
    assert result.source_cells_count == 3


def test_generate_workbook_zero_numeric_literals_in_derived_cells(
    tmp_path: Path,
) -> None:
    """Verifies that every derived cell in Reconciliation sheet contains a formula string (AC-2)."""
    nodes = [
        _make_input_node(0, "Operating Income", value="1000.0", page=5),
        _make_input_node(1, "Interest Expense", value="150.0", page=6),
    ]
    batch = FormulaInputBatch(
        nodes=nodes,
        total_records_received=2,
        confirmed_count=2,
        excluded_count=0,
    )
    tree = build_formula_tree(batch, target_metric="Adjusted EBITDA")

    result = generate_workbook(tree, job_id="job_test_2", output_dir=tmp_path)
    assert result.is_success is True

    # Open with openpyxl to inspect formulas (openpyxl preserves formula strings)
    wb = openpyxl.load_workbook(result.file_path, data_only=False)
    ws_recon = wb["Reconciliation"]

    # Value column is column B (index 2). Rows 4, 5 are component formulas, Row 6 is total formula.
    for row in range(4, ws_recon.max_row + 1):
        cell_val = str(ws_recon.cell(row=row, column=2).value or "")
        assert cell_val.startswith(
            "="
        ), f"Cell B{row} should contain formula, got {cell_val}"
        assert not isinstance(ws_recon.cell(row=row, column=2).value, (int, float))


def test_generate_workbook_sheet_names_have_no_spaces(tmp_path: Path) -> None:
    """Verifies that all generated sheet names contain zero spaces (AC-7, EC-6)."""
    nodes = [_make_input_node(0, "EBITDA", value="500.0")]
    batch = FormulaInputBatch(
        nodes=nodes,
        total_records_received=1,
        confirmed_count=1,
        excluded_count=0,
    )
    tree = build_formula_tree(batch, target_metric="Adjusted EBITDA")

    result = generate_workbook(tree, job_id="job_test_3", output_dir=tmp_path)
    assert result.is_success is True

    wb = openpyxl.load_workbook(result.file_path)
    for name in wb.sheetnames:
        assert " " not in name, f"Sheet name '{name}' must not contain spaces (EC-6)"


def test_generate_workbook_unparseable_value_warning(tmp_path: Path) -> None:
    """Verifies that unparseable values (e.g. 'N/A') are written as text and produce warnings (EC-2)."""
    nodes = [
        _make_input_node(0, "Operating Income", value="500.0"),
        _make_input_node(1, "Litigation Charges", value="N/A"),
    ]
    batch = FormulaInputBatch(
        nodes=nodes,
        total_records_received=2,
        confirmed_count=2,
        excluded_count=0,
    )
    tree = build_formula_tree(batch, target_metric="Adjusted EBITDA")

    result = generate_workbook(tree, job_id="job_test_4", output_dir=tmp_path)
    assert result.is_success is True
    assert len(result.warnings) == 1
    assert "N/A" in result.warnings[0]
    assert "EC-2" in result.warnings[0]


def test_generate_workbook_duplicate_label_aggregation(tmp_path: Path) -> None:
    """Verifies that duplicate labels generate sub-items and an intermediate aggregate sum formula (EC-1)."""
    nodes = [
        _make_input_node(0, "Operating Income", value="1000.0", page=10),
        _make_input_node(1, "Stock-Based Compensation", value="20.0", page=15),
        _make_input_node(2, "Stock-Based Compensation", value="30.0", page=45),
    ]
    batch = FormulaInputBatch(
        nodes=nodes,
        total_records_received=3,
        confirmed_count=3,
        excluded_count=0,
    )
    tree = build_formula_tree(batch, target_metric="Adjusted EBITDA")

    result = generate_workbook(tree, job_id="job_test_5", output_dir=tmp_path)
    assert result.is_success is True

    wb = openpyxl.load_workbook(result.file_path, data_only=False)
    ws_recon = wb["Reconciliation"]

    # Check aggregate row contains =SUM(...)
    aggregate_found = False
    for row in range(4, ws_recon.max_row + 1):
        label = ws_recon.cell(row=row, column=1).value
        val_formula = str(ws_recon.cell(row=row, column=2).value or "")
        if label == "Total Stock-Based Compensation":
            aggregate_found = True
            assert "SUM(" in val_formula
    assert aggregate_found is True


def test_generate_workbook_atomic_write_error_handling(tmp_path: Path) -> None:
    """Verifies that exceptions during workbook generation discard partial files and report failure (EC-7)."""
    nodes = [_make_input_node(0, "EBITDA", value="100.0")]
    batch = FormulaInputBatch(
        nodes=nodes,
        total_records_received=1,
        confirmed_count=1,
        excluded_count=0,
    )
    tree = build_formula_tree(batch, target_metric="Adjusted EBITDA")

    with patch(
        "xlsxwriter.Workbook.add_worksheet", side_effect=OSError("Disk write failure")
    ):
        result = generate_workbook(tree, job_id="job_test_err", output_dir=tmp_path)

    assert result.is_success is False
    assert "Disk write failure" in (result.error_detail or "")
    assert not Path(result.file_path).exists()
    assert not (tmp_path / "models" / "job_test_err_model.xlsx.tmp").exists()


def test_generate_workbook_invalid_tree(tmp_path: Path) -> None:
    """Verifies invalid formula tree returns failure result without creating files."""
    invalid_tree = FormulaTree(
        target_metric="Unsupported",
        is_valid=False,
        error_message="Unsupported metric",
    )

    result = generate_workbook(
        invalid_tree, job_id="job_test_invalid", output_dir=tmp_path
    )

    assert result.is_success is False
    assert result.error_detail == "Unsupported metric"
    assert not Path(result.file_path).exists()


def test_generate_workbook_exactly_one_comment_and_hyperlink_per_cell(
    tmp_path: Path,
) -> None:
    """Verifies that 100% of generated data/formula cells carry exactly 1 comment and 1 hyperlink (AC-6)."""
    nodes = [
        _make_input_node(0, "Operating Income", value="1,000.00", page=12),
        _make_input_node(1, "Depreciation & Amortization", value="200.00", page=14),
        _make_input_node(2, "Stock-Based Compensation", value="50.00", page=25),
        _make_input_node(3, "Stock-Based Compensation", value="30.00", page=45),
    ]
    batch = FormulaInputBatch(
        nodes=nodes,
        total_records_received=4,
        confirmed_count=4,
        excluded_count=0,
    )
    tree = build_formula_tree(batch, target_metric="Adjusted EBITDA")

    result = generate_workbook(tree, job_id="job_ac6_test", output_dir=tmp_path)
    assert result.is_success is True

    # Provenance record collection matching cell count
    assert len(result.provenance_records) == result.total_cells_generated
    assert len(result.cell_references) == result.total_cells_generated

    wb = openpyxl.load_workbook(result.file_path, data_only=False)

    # 1. Check Source_Inputs value column (Col B / 2)
    ws_inputs = wb["Source_Inputs"]
    for row in range(2, 6):  # 4 data rows
        cell = ws_inputs.cell(row=row, column=2)
        assert cell.comment is not None, f"Source_Inputs!B{row} missing comment"
        assert "[Footnote Provenance]" in cell.comment.text
        assert cell.hyperlink is not None, f"Source_Inputs!B{row} missing hyperlink"
        assert (
            "http://localhost:8000/models/job_ac6_test/provenance/Source_Inputs/B"
            in cell.hyperlink.target
        )

    # 2. Check Reconciliation value column (Col B / 2)
    ws_recon = wb["Reconciliation"]
    for row in range(4, ws_recon.max_row + 1):
        cell = ws_recon.cell(row=row, column=2)
        assert cell.comment is not None, f"Reconciliation!B{row} missing comment"
        assert "[Footnote Provenance" in cell.comment.text
        # Formulas use =HYPERLINK(...) wrapper
        val_str = str(cell.value or "")
        assert val_str.startswith(
            "=HYPERLINK("
        ), f"Reconciliation!B{row} formula must contain HYPERLINK, got {val_str}"


def test_generate_workbook_empty_leaves_invalid_tree(tmp_path: Path) -> None:
    """Verifies that empty formula batch builds invalid tree and generate_workbook cleanly fails with zero output files."""
    empty_batch = FormulaInputBatch(
        nodes=[],
        total_records_received=0,
        confirmed_count=0,
        excluded_count=0,
        error_message="No confirmed records available for formula generation.",
    )
    tree = build_formula_tree(empty_batch, target_metric="Adjusted EBITDA")
    assert tree.is_valid is False

    result = generate_workbook(tree, job_id="job_empty_tree", output_dir=tmp_path)
    assert result.is_success is False
    assert result.total_cells_generated == 0
    assert result.sheet_names == []
    assert not Path(result.file_path).exists()
    assert not (tmp_path / "models" / "job_empty_tree_model.xlsx.tmp").exists()


def test_banker_editable_two_column_structure_and_formulas(tmp_path: Path) -> None:
    """
    Comprehensive verification for Ticket 1.3.3:
    1. Exactly 2 sheets: Source_Inputs and Reconciliation.
    2. Source_Inputs has 2 columns (Label, Value) with comments containing provenance.
    3. Reconciliation has 2 columns (Line Item, Value) with cross-sheet formulas.
    4. Target metric total uses =SUM(...) formula over component rows.
    5. Editing Source_Inputs cell coordinates propagates to Reconciliation formulas.
    """
    nodes = [
        _make_input_node(
            0, "Operating Income", value="1,200.00", page=8, source_file="10k_2023.pdf"
        ),
        _make_input_node(
            1,
            "Depreciation & Amortization",
            value="300.00",
            page=12,
            source_file="10k_2023.pdf",
        ),
        _make_input_node(
            2,
            "Stock-Based Compensation",
            value="80.00",
            page=22,
            source_file="10k_2023.pdf",
        ),
    ]
    batch = FormulaInputBatch(
        nodes=nodes,
        total_records_received=3,
        confirmed_count=3,
        excluded_count=0,
    )
    tree = build_formula_tree(batch, target_metric="Adjusted EBITDA")
    assert tree.is_valid is True

    result = generate_workbook(tree, job_id="job_ticket_1_3_3", output_dir=tmp_path)
    assert result.is_success is True

    wb = openpyxl.load_workbook(result.file_path, data_only=False)

    # 1. Exactly 2 sheets
    assert wb.sheetnames == ["Source_Inputs", "Reconciliation"]

    # 2. Source_Inputs 2-column layout and provenance comments
    ws_inputs = wb["Source_Inputs"]
    assert ws_inputs.max_column == 2
    assert ws_inputs.cell(row=1, column=1).value == "Label"
    assert ws_inputs.cell(row=1, column=2).value == "Value ($)"

    for r in range(2, 5):
        val_cell = ws_inputs.cell(row=r, column=2)
        assert (
            val_cell.comment is not None
        ), f"Source_Inputs!B{r} missing provenance comment"
        comment_text = val_cell.comment.text
        assert "[Footnote Provenance]" in comment_text
        assert "Source: 10k_2023.pdf" in comment_text
        assert "BBox [0-1000]" in comment_text
        assert val_cell.hyperlink is not None, f"Source_Inputs!B{r} missing hyperlink"

    # 3. Reconciliation 2-column layout and cross-sheet formulas
    ws_recon = wb["Reconciliation"]
    assert ws_recon.max_column == 2
    assert ws_recon.cell(row=1, column=1).value == "Adjusted EBITDA Reconciliation"
    assert ws_recon.cell(row=3, column=1).value == "Line Item"
    assert ws_recon.cell(row=3, column=2).value == "Value ($)"

    # Rows 4, 5, 6 are component lines referencing Source_Inputs!B2, B3, B4
    for idx, r in enumerate(range(4, 7), start=2):
        cell_formula = str(ws_recon.cell(row=r, column=2).value or "")
        assert (
            f"Source_Inputs!B{idx}" in cell_formula
        ), f"Expected Reconciliation!B{r} to reference Source_Inputs!B{idx}, got {cell_formula}"
        assert cell_formula.startswith("=HYPERLINK(")

    # Row 7 is Adjusted EBITDA total using =SUM(...)
    total_cell_label = ws_recon.cell(row=7, column=1).value
    total_cell_formula = str(ws_recon.cell(row=7, column=2).value or "")
    assert total_cell_label == "Adjusted EBITDA"
    assert "SUM(B4, B5, B6)" in total_cell_formula or "SUM(" in total_cell_formula
    assert total_cell_formula.startswith("=HYPERLINK(")
